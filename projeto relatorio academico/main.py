import openpyxl

try:
    workbook_original = openpyxl.load_workbook("alunos projeto.xlsx")
    planilha_alunos = workbook_original.active
except FileNotFoundError:
    print("Erro: O arquivo 'alunos projeto.xlsx' não foi encontrado na pasta.")
    exit()

dados_aprovados = []
dados_reprovados = []

soma_notas = 0
total_alunos = 0
maior_nota = -1
aluno_destaque = ""

cabecalhos = list(next(planilha_alunos.iter_rows(min_row=1, max_row=1, values_only=True)))

for linha in planilha_alunos.iter_rows(min_row=2, values_only=True):

    nome, curso, idade, nota_final, data_matricula = linha
    
    if nome is None:
        continue
  
    total_alunos += 1
    soma_notas += nota_final
    
    if nota_final > maior_nota:
        maior_nota = nota_final
        aluno_destaque = nome

    if nota_final >= 7.0:
        dados_aprovados.append(linha)
    else:
        dados_reprovados.append(linha)

wb_aprovados = openpyxl.Workbook()
planilha_aprovados = wb_aprovados.active
planilha_aprovados.title = "Aprovados"

planilha_aprovados.append(cabecalhos)
for aluno in dados_aprovados:
    planilha_aprovados.append(aluno)
    
wb_aprovados.save("aprovados.xlsx")

wb_reprovados = openpyxl.Workbook()
planilha_reprovados = wb_reprovados.active
planilha_reprovados.title = "Reprovados"

planilha_reprovados.append(cabecalhos)
for aluno in dados_reprovados:
    planilha_reprovados.append(aluno)
    
wb_reprovados.save("reprovados.xlsx")

media_turma = soma_notas / total_alunos if total_alunos > 0 else 0

print("\n" + "="*35)
print(" RELATÓRIO ACADÊMICO GERAL")
print("="*35)
print(f"Total de Aprovados: {len(dados_aprovados)}")
print(f"Total de Reprovados: {len(dados_reprovados)}")
print(f"Nota Média da Turma: {media_turma:.2f}")
print(f"Aluno com Maior Nota: {aluno_destaque} (Nota: {maior_nota:.2f})")
print("="*35)
print("Arquivos 'aprovados.xlsx' e 'reprovados.xlsx' gerados com sucesso!\n")